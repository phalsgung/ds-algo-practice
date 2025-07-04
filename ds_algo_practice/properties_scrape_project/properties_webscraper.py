import time
import random
import re

from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
import requests
from selenium import webdriver
from selenium.webdriver import DesiredCapabilities


class ZillowScraper:

    def __init__(self):
        self.driver = None
        capabilities_chrome = DesiredCapabilities().CHROME
        capabilities_chrome["pageLoadStrategy"] = "eager"
        self.driver = webdriver.Chrome()

    def get_scraped_data(self, address, city, state, zipcode):
        if pd.isna(address) or pd.isna(city) or pd.isna(state) or pd.isna(zipcode):
            raise Exception("Invalid Address")

        parsed_data = {
            'zestimate': '',
            'rent_zestimate': '',
        }

        # Ensure address is a string before replacing spaces
        address = str(address).replace(" ", "-")
        city = str(city).replace(" ", "-")
        state = str(state).replace(" ", "-")
        zipcode = str(int(zipcode))  # Convert to string and remove decimal point if it's a float

        # Construct the Zillow URL for the property
        url = f"https://www.zillow.com/homes/{address},-{city},-{state}-{zipcode}/"
        print("Scraping URL: ", url)

        # response = requests_session.get(url, headers=headers)
        self.driver.get(url)
        # print("Response:\n", self.driver.page_source)

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        # zestimate
        try:
            # span: data-testid="zestimate-text" > span > innertext
            zestimate_field_grp = soup.select('span[data-testid="zestimate-text"]')[0]
            parsed_data['zestimate'] = zestimate_field_grp.find('span').get_text()
        except Exception:
            try:
                # 2nd Attempt
                parsed_data['zestimate'] = soup.select('p[data-testid="primary-zestimate"]')[0].get_text()
            except Exception as exc:
                print(f"Unknown Exception in getting zestimate: {exc}")
                parsed_data['zestimate'] = 'Zestimate not found'
        # rent_zestimate
        try:
            rent_zestimate_field_grp = soup.select('span[data-testid="zestimate-text"]')[1]
            rent_zestimate = rent_zestimate_field_grp.find('span').get_text()
            parsed_data['rent_zestimate'] = rent_zestimate
        except Exception:
            try:
                # 2nd Attempt
                parsed_data['rent_zestimate'] = soup.select('p[data-testid="rent-zestimate"]')[0].get_text()
            except Exception as exc:
                print(f"Unknown Exception in getting rent_zestimate: {exc}")
                parsed_data['rent_zestimate'] = 'Rent Zestimate not found'

        return parsed_data

    def close(self):
        if self.driver is not None:
            self.driver.close()


class RedfinScraper:

    def __init__(self):
        self.driver = None
        capabilities_chrome = DesiredCapabilities().CHROME
        capabilities_chrome["pageLoadStrategy"] = "eager"
        self.driver = webdriver.Chrome()

    def _get_address_url(self, address, city, state, zipcode):
        # Ensure address is a string before replacing spaces
        address = str(address).replace(" ", "%20")
        city = str(city).replace(" ", "%20")
        state = str(state).replace(" ", "%20")
        zipcode = str(int(zipcode))  # Convert to string and remove decimal point if it's a float

        location_string = f"{address}%20{city}%20{state}%20{zipcode}"
        url = f"https://www.redfin.com/stingray/do/query-location?location={location_string}&v=2"
        self.driver.get(url)

        regex = r"\"urlV2\":\"([A-Za-z0-9-\/]*)\","
        regexp_1 = re.compile(regex)
        re_match = regexp_1.search(self.driver.page_source, re.MULTILINE)
        time.sleep(random.randint(2, 5))
        if re_match:
            urlv2 = re_match.group(1)
            return urlv2
        else:
            print(f"ERROR: Could not find URL for location_string: {location_string} from response_text: {self.driver.page_source}")
            return None

    def get_scraped_data(self, address, city, state, zipcode):
        if pd.isna(address) or pd.isna(city) or pd.isna(state) or pd.isna(zipcode):
            raise Exception("Invalid Address")

        parsed_data = {
            'redfin_estimate': '',
        }

        property_url = self._get_address_url(address, city, state, zipcode)
        if property_url is None:
            return parsed_data

        # Construct the Redfin URL for the property
        url = f"https://www.redfin.com{property_url}"
        print("Scraping URL: ", url)

        # response = requests_session.get(url, headers=headers)
        self.driver.get(url)
        # print("Response:\n", self.driver.page_source)

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        # redfin_estimate
        try:
            redfin_estimate_div_grp = soup.select('div[data-rf-test-id="abp-price"]')[0]
            parsed_data['redfin_estimate'] = redfin_estimate_div_grp.find('div').get_text()
        except Exception as exc:
            print(f"Unknown Exception in getting redfin_estimate: {exc}")
            parsed_data['redfin_estimate'] = 'Redfin Estimate not found'

        return parsed_data

    def close(self):
        if self.driver is not None:
            self.driver.close()


def scrap_properties():
    print("Properties Scraper Started")

    try:
        # Load the Excel file
        file_path = "./Homes details_Template.xlsx"
        input_excel_df = pd.read_excel(file_path)
        input_excel_wb = openpyxl.load_workbook(file_path)
        input_excel_ds = input_excel_wb["Home Details"]
        print(input_excel_df)
    except Exception as exc:
        print("Unknown Error in opening Input Excel")
        raise(exc)

    try:
        # zillow_scraper = ZillowScraper()
        redfin_scraper = RedfinScraper()
        pass
    except Exception as exc:
        print("Unknown Exception in Starting Properties Scraper(s)")
        raise(exc)

    # Iterate through the Excel data
    for row_index, row in input_excel_df.iterrows():
        try:
            property_number = int(row['Property #'])
        except Exception as exc:
            continue

        address = row['Street address']
        city = row['City']
        state = row['State']
        zipcode = row['Zip']

        try:
            parsed_data = {
                'zestimate': '',
                'rent_zestimate': '',
                'redfin_estimate': '',
            }
            if pd.notna(address) and pd.notna(city) and pd.notna(state) and pd.notna(zipcode):
                # parsed_data.update(zillow_scraper.get_scraped_data(address, city, state, zipcode))
                parsed_data.update(redfin_scraper.get_scraped_data(address, city, state, zipcode))
                pass
            else:
                parsed_data = {
                    'zestimate': 'Invalid Address',
                    'rent_zestimate': 'Invalid Address',
                    'redfin_estimate': 'Invalid Address',
                }
        except Exception as exc:
            print("Unknown Exception in Getting Properties Data")
            parsed_data = {
                'zestimate': f"Exception: {exc}",
                'rent_zestimate': f"Exception: {exc}",
                'redfin_estimate': f"Exception: {exc}",
            }

        input_excel_ds.cell(row=row_index+2, column=input_excel_df.columns.get_loc("Zestimate")+1).value = str(parsed_data['zestimate'])  # Explicitly cast Zestimate to string to avoid dtype issues
        input_excel_ds.cell(row=row_index+2, column=input_excel_df.columns.get_loc("Rent Zestimate")+1).value = str(parsed_data['rent_zestimate'])  # Explicitly cast Zestimate to string to avoid dtype issues
        input_excel_ds.cell(row=row_index+2, column=input_excel_df.columns.get_loc("Redfin Estimate")+1).value = str(parsed_data['redfin_estimate'])  # Explicitly cast Zestimate to string to avoid dtype issues

        # Print the Properties Details in the terminal
        print(f"Scraped Details for {address}, {city}, {state} {zipcode}: {parsed_data}")

        # Add a random delay between requests to avoid being blocked
        time.sleep(random.randint(2, 5))

    # zillow_scraper.close()
    redfin_scraper.close()

    # Save updated data to the original Excel file
    input_excel_wb.save(file_path)

    print("Scraped data updated successfully.")
    print("Properties Scraper Ended")


def main():
    scrap_properties()


if __name__=="__main__":
    main()
